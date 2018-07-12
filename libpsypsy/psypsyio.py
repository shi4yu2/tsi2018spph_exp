#!/usr/bin/env python3
# -*- coding: utf-8 -*-

"""
PSYchology Python by Shi Yu
I/O handler
"""

# Todo :
#

__author__ = 'ShY'
__copyright__ = 'Copyright 2018, SHY'
__version__ = '0.2.0 (20180531)'
__maintainer__ = 'ShY, Pierre Halle'
__email__ = 'shi4yu2@gmail.com'
__status__ = 'Development'

import re
from openpyxl import load_workbook
from pathlib import Path
import csv


# ====================== Input =========================
# ======================================================
# Processing text files
# This function returns a dictionary of stimuli
# containing keys:
#     header, header_length, conditions, trial_number
def processing_stimuli_file(filename, separator='\t', header=True):
    # type: (str, str, bool) -> tuple
    """
    :param filename: input file name
    :type: filename: str
    :param separator: separator of text file
    :type: separator: str
    :param header: if header line exists
    :type: bool
    :return: output_stimuli_dictionary: dictionary of stimuli
    :rtype: tuple
    """
    f = open(filename, mode='rU', encoding="utf8")
    plain_text = f.read()
    line_buffer = cleanup_line_breaks(plain_text, regex='[\n|\r]+')
    f.close()

    # Handle header
    list_header = filter_empty_item(line_buffer[0].split(separator))
    header_length = len(list_header)

    if not header:  # when file does not contain a header
        list_header = list(range(header_length))

    # Create a dictionary containing a list of header items and the length of header
    output_stimuli_dictionary = {'header': list_header, 'header_length': header_length}

    # Create header conditions in output dictionary
    header_index = {}
    for condition in list_header:
        index = list_header.index(condition)
        header_index[index] = condition
        output_stimuli_dictionary[condition] = []

    # Processing list of stimuli
    if header:
        start_line = 1
    else:
        start_line = 0

    # Fill the dictionary of stimuli
    for line in line_buffer[start_line:]:
        list_line = filter_empty_item(line.split('\t'))
        for item in range(header_length):
            output_stimuli_dictionary[header_index[item]].append(list_line[item])

    # Compute number of trials
    trial_number = len(output_stimuli_dictionary[list_header[1]])
    output_stimuli_dictionary['trial_number'] = trial_number

    return output_stimuli_dictionary, header_index


# Filter special line break characters =================
def cleanup_line_breaks(text, regex='[\n|\r]+'):
    # type: (str, str) -> list
    """
    :param text: string of content of input file
    :type: text: str
    :param regex: regular expression for line breaks
    :type: regex: str
    :return: line_buffer: list of strings
    :rtype: list
    """
    re_line_break = re.compile(regex)  # use regex to filter line breaks
    line_buffer = re_line_break.split(text)  # chunk text into a list of lines
    line_buffer = filter_empty_item(line_buffer)  # remove empty lines

    return line_buffer


# Filter empty items in a list
def filter_empty_item(list_origin):
    # type: (list) -> list
    """
    :param list_origin: original list containing empty items
    :type: list_origin: list
    :return: list without empty items
    """

    return list(filter(None, list_origin))


# Read excel files and returns a dictionary of stimuli
# containing keys:
#     header, header_length, conditions, trial_number
def processing_stimuli_file_excel(filename):
    # type: (str) -> tuple
    """
    Processing excel_file input
    :param filename: Excel file
    :type filename: str
    :returns: dictionary of stimuli
    :rtype: dict
    """
    output_stimuli_dictionary = {}
    title_sheet = []
    wb = load_workbook(filename=filename, read_only=True)

    # header
    first_sheet = wb.get_sheet_names()[0]
    ws = wb.get_sheet_by_name(first_sheet)
    first_row_ws = ws[1]
    for item in first_row_ws:
        title_sheet.append(item.value)
    header_length = len(title_sheet)

    header_index = {}
    for condition in title_sheet:
        index = title_sheet.index(condition)
        header_index[index] = condition

    output_stimuli_dictionary['header_length'] = header_length  # get header_length
    output_stimuli_dictionary['header'] = title_sheet

    # stimuli
    # empty dictionary with conditions
    for condition in title_sheet:
        output_stimuli_dictionary[str(condition)] = []

    for row in ws.iter_rows(row_offset=1):
        for cell in row:
            file_column = cell.column
            title_column = ws.cell(row=1, column=file_column).value
            output_stimuli_dictionary[title_column].append(cell.value.decode('utf-8'))

    # get number of trials
    trial_number = ws.max_row - 1
    output_stimuli_dictionary['trial_number'] = trial_number

    return output_stimuli_dictionary, header_index


def read_stimuli(filename, separator='\t', header=True):
    # type: (str, str, bool) -> tuple
    """
    Processing stimuli input
    :param filename: file (text or excel)
    :type filename: str
    :param separator: separator of column
    :type separator: str
    :param header: if header line exists
    :type header: bool
    :returns: dictionary of stimuli
    :rtype: Union[dict, tuple]
    """
    if filename.endswith('.xlsx'):
        return processing_stimuli_file_excel(filename)
    else:
        return processing_stimuli_file(filename, separator, header)

# ====================== Input =========================


# ====================== Output ========================
# ======================================================
def write_result_line(f, result_line, separator="\t"):
    # type: (TextIO, list, str) -> ()
    """
    Write result into file
    :param f: result file
    :type: f: textIO
    :param result_line: result table
    :type: result_line: list
    :param separator: column separator
    :type: separator: str
    """
    for i in range(len(result_line)):
        if i == len(result_line) - 1:
            print((result_line[i]), end="\n", file=f)
        else:
            print((result_line[i]), end=separator, file=f)
    return


def write_result_table(filename, result, separator="\t"):
    # type: (str, list, str) -> ()
    """
    Write result into file
    :param filename: result file
    :type: filename: str
    :param result result table
    :type: result: list
    :param separator: column separator
    :type: separator: str
    """
    f = open(filename, 'w')
    for result_line in result:
        for i in range(len(result_line)):
            if i == len(result_line) - 1:
                print((result_line[i]), end="\n", file=f)
            else:
                print((result_line[i]), end=separator, file=f)
    f.close()
    return


def write_result_header(filename, trial, result_columns):
    # type: (TextIO, dict, list) -> ()
    """
    Write result header
    :param filename: result filename
    :type filename: textIO
    :param trial: dictionary of all trials
    :type trial: dict
    :param result_columns: additional column for results
    :type result_columns: list
    :return:
    """
    # Result header
    result_header = trial["header"][:]
    result_header.extend(result_columns)
    # print(result_header)
    write_result_line(filename, result_header)
    return


def read_csv(filename):
    # type: (str) -> tuple[list, list]
    """
    Read CSV files and return a list of lists
    :param filename: filename
    :type filename: str
    :return: list of lists of str, list of header
    :rtype: tuple
    """
    try:
        with open(filename, 'r') as csv_file:
            dialect = csv.Sniffer().sniff(csv_file.readline(), [',', ';', '\t', ' '])
            csv_file.seek(0)
            data_frame = csv.reader(csv_file, dialect)
            output_list = []
            for row_i in data_frame:
                output_list.append(row_i)
            header = output_list.pop(0)
            return output_list, header
    except OSError:
        print("File not found")


def write_total_result(filename, subj_n, result_input, separator='\t'):
    result_file = Path(filename)
    result, header = read_csv(result_input)

    first = result_file.is_file()
    f = open(filename, 'a')

    if not first:
        header.append("subj")
        write_result_line(f, header)

    line_length = len(result[0]) + 1

    for i in range(len(result)):
        result[i] = result[i][:] + [subj_n]
        for j in range(line_length):
            if j == line_length - 1:
                print((result[i][j]), end="\n", file=f)
            else:
                print((result[i][j]), end=separator, file=f)
    f.close()
    return
